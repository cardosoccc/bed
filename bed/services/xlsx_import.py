import re
from datetime import date, datetime

import openpyxl

INSTITUTION_MAP = {
    "INTER DISTRIBUIDORA DE TITULOS E VALORES MOBILIARIOS LTDA": "inter",
    "INTER DTVM LTDA": "inter",
    "BANCO BTG PACTUAL S/A": "btg-pactual",
}

ACCENT_NORMALIZATION = {
    "Transferencia": "Transferência",
}

NO_TICKER_PREFIXES = ("Tesouro ", "CDB ")

AGF_TYPE_NORMALIZATION = {
    "Jscp": "Juros Sobre Capital Próprio",
    "RENDIMENTO": "Rendimento",
}

MOV_TYPE_WHITELIST = {
    "Dividendo",
    "Juros Sobre Capital Próprio",
    "Rendimento",
    "Reembolso",
    "Compra",
    "Venda",
    "COMPRA / VENDA",
    "VENCIMENTO",
}


def parse_xlsx(file_path: str, sheet_name: str = "Movimentação") -> list[dict]:
    """Parse B3 movimentação XLSX into a list of raw transaction dicts."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        entrada_saida, data, movimentacao, produto, instituicao, quantidade, preco, valor = row[:8]

        rows.append({
            "entrada_saida": str(entrada_saida).strip(),
            "date": parse_date(data),
            "type": normalize_type(str(movimentacao).strip()),
            "product": str(produto).strip(),
            "ticker": extract_ticker(str(produto).strip()),
            "institution": map_institution(str(instituicao).strip()),
            "quantity": parse_numeric(quantidade),
            "unit_value": parse_numeric(preco),
            "total_value": parse_numeric(valor),
        })

    wb.close()
    return rows


def parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()


def parse_numeric(value) -> float:
    if value is None or str(value).strip() in ("-", ""):
        return 0.0
    return float(value)


def extract_ticker(product: str) -> str | None:
    for prefix in NO_TICKER_PREFIXES:
        if product.startswith(prefix):
            return None
    if " - " in product:
        return product.split(" - ")[0].strip()
    return None


def map_institution(name: str) -> str:
    if name in INSTITUTION_MAP:
        return INSTITUTION_MAP[name]
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return slug


def normalize_type(movimentacao: str) -> str:
    return ACCENT_NORMALIZATION.get(movimentacao, movimentacao)


def apply_sign(total_value: float, entrada_saida: str) -> float:
    if entrada_saida == "Debito":
        return -abs(total_value)
    return abs(total_value)


def normalize_agf_type(tipo: str) -> str:
    return AGF_TYPE_NORMALIZATION.get(tipo, tipo)


def strip_fracionario_suffix(ticker: str) -> str:
    if re.match(r"^[A-Z]{4}\d+F$", ticker):
        return ticker[:-1]
    return ticker


def parse_mov_xlsx(file_path: str, sheet_name: str = "Movimentação") -> list[dict]:
    """Parse B3 movimentação XLSX, filtering to whitelisted types only."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        entrada_saida, data, movimentacao, produto, instituicao, quantidade, preco, valor = row[:8]
        tipo = normalize_type(str(movimentacao).strip())

        if tipo not in MOV_TYPE_WHITELIST:
            continue

        rows.append({
            "entrada_saida": str(entrada_saida).strip(),
            "date": parse_date(data),
            "type": tipo,
            "product": str(produto).strip(),
            "ticker": extract_ticker(str(produto).strip()),
            "institution": map_institution(str(instituicao).strip()),
            "quantity": parse_numeric(quantidade),
            "unit_value": parse_numeric(preco),
            "total_value": parse_numeric(valor),
        })

    wb.close()
    return rows


def parse_agf_xlsx(file_path: str, institution: str = "inter") -> list[dict]:
    """Parse AGF+ XLSX export into a list of raw transaction dicts."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0] or str(row[0]).strip() == "Data Operação":
            continue

        data, empresa, num_acoes, tipo, _origem, valor, _irrf, total_liquido = row[:8]

        product = str(empresa).strip() if empresa else ""
        ticker = extract_ticker(product)
        quantity = abs(parse_numeric(num_acoes))
        unit_value = parse_numeric(valor)
        total_value = parse_numeric(total_liquido)

        normalized_type = normalize_agf_type(str(tipo).strip())

        if normalized_type == "Venda":
            total_value = -abs(total_value)

        rows.append({
            "date": parse_date(data),
            "type": normalized_type,
            "product": product,
            "ticker": ticker,
            "institution": institution,
            "quantity": quantity,
            "unit_value": unit_value,
            "total_value": total_value,
        })

    wb.close()
    return rows


def parse_neg_xlsx(file_path: str) -> list[dict]:
    """Parse B3 negociação XLSX export into a list of raw transaction dicts."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        data, tipo, _mercado, _prazo, instituicao, codigo, quantidade, preco, valor = row[:9]

        ticker = strip_fracionario_suffix(str(codigo).strip())
        tipo_str = str(tipo).strip()
        total_value = parse_numeric(valor)

        if tipo_str == "Venda":
            total_value = -abs(total_value)

        rows.append({
            "date": parse_date(data),
            "type": tipo_str,
            "product": ticker,
            "ticker": ticker,
            "institution": map_institution(str(instituicao).strip()),
            "quantity": parse_numeric(quantidade),
            "unit_value": parse_numeric(preco),
            "total_value": total_value,
        })

    wb.close()
    return rows
